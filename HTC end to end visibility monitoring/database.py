import sqlite3
import openpyxl
import re
import os
import logging
from datetime import datetime

logger = logging.getLogger(__name__)

def init_db(db_path):
    """Initializes the database schema."""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS htc_events (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        assembly_cd TEXT,
        nh_assembly_cd TEXT,
        bom_class_cd TEXT,
        config_slot_code TEXT,
        config_slot_name TEXT,
        part_group_cd TEXT,
        part_group_name TEXT,
        event_dt TEXT,
        event_type TEXT,
        status_cd TEXT,
        event_desc TEXT,
        aircraft TEXT,
        inventory_key TEXT,
        inventory TEXT,
        barcode TEXT,
        config_slot TEXT,
        part_no TEXT,
        part_desc TEXT,
        remove_reason TEXT,
        performed_by_user TEXT,
        performed_by_username TEXT,
        performed_by_hr_cd TEXT,
        serial_number TEXT,
        has_xxx_sn INTEGER,
        UNIQUE(barcode, event_dt, event_type, config_slot_code, inventory_key)
    )
    ''')
    
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS resolutions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        event_barcode TEXT,
        alert_date TEXT,
        aircraft TEXT,
        config_slot TEXT,
        part_no TEXT,
        original_sn TEXT,
        resolved_sn TEXT,
        engineer_responsible TEXT,
        resolution_date TEXT,
        status TEXT DEFAULT 'PENDING',
        notes TEXT,
        created_at TEXT,
        updated_at TEXT
    )
    ''')

    cursor.execute('''
    CREATE TABLE IF NOT EXISTS loaded_files (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        filename TEXT UNIQUE,
        filepath TEXT,
        records_loaded INTEGER DEFAULT 0,
        loaded_at TEXT,
        source TEXT DEFAULT 'manual'
    )
    ''')

    # ── Email subscriptions (HTC Visibility Monitoring digest emails) ────────
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS email_subscriptions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fleet_type TEXT NOT NULL,
        alert_type TEXT NOT NULL,           -- comma-separated list, e.g. 'ALERT_B,ALERT_C'
        frequency TEXT NOT NULL,            -- 'every_3_days' or 'weekly'
        day_of_week TEXT,                   -- 'mon'..'sun', required if frequency='weekly'
        run_time TEXT NOT NULL,             -- 'HH:MM' 24h, local server time
        email TEXT NOT NULL,
        active INTEGER DEFAULT 1,
        created_at TEXT,
        updated_at TEXT,
        last_sent_at TEXT
    )
    ''')

    # ── Generic key/value app config (OWA password, admin password, etc.) ────
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS app_config (
        key TEXT PRIMARY KEY,
        value TEXT
    )
    ''')

    conn.commit()
    conn.close()


def is_file_loaded(db_path, filename):
    """Check if a file has already been imported."""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM loaded_files WHERE filename = ?", (filename,))
    count = cursor.fetchone()[0]
    conn.close()
    return count > 0


def mark_file_loaded(db_path, filename, filepath, records_loaded, source='manual'):
    """Record that a file has been imported."""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    now = datetime.utcnow().isoformat()
    try:
        cursor.execute('''
        INSERT OR REPLACE INTO loaded_files (filename, filepath, records_loaded, loaded_at, source)
        VALUES (?, ?, ?, ?, ?)
        ''', (filename, filepath, records_loaded, now, source))
        conn.commit()
    except Exception as e:
        logger.error(f"Error marking file as loaded: {e}")
    finally:
        conn.close()


def parse_filename_timestamp(filename, filepath):
    """Parse timestamps from filenames."""
    # Pattern 1: MON-DD-YYYY (e.g., JUN-17-2026)
    match = re.search(r'([A-Za-z]{3})-(\d{2})-(\d{4})', filename)
    if match:
        try:
            return datetime.strptime(match.group(0).upper(), '%b-%d-%Y')
        except ValueError:
            pass
            
    # Pattern 2: YYYYMMDDHHMI (e.g., 202607200837)
    match = re.search(r'(\d{12})', filename)
    if match:
        try:
            return datetime.strptime(match.group(1), '%Y%m%d%H%M')
        except ValueError:
            pass
            
    # Pattern 3: YYYY-MM-DD (e.g., 2026-07-17)
    match = re.search(r'(\d{4}-\d{2}-\d{2})', filename)
    if match:
        try:
            return datetime.strptime(match.group(1), '%Y-%m-%d')
        except ValueError:
            pass
            
    # Fallback
    try:
        return datetime.fromtimestamp(os.path.getmtime(filepath))
    except Exception:
        return datetime.min

def scan_share_directory(share_path):
    """Scan a network share or local directory for .xlsx and .xls files.
    Returns a list of (filename, full_path, parsed_timestamp) tuples, sorted latest first.
    """
    results = []
    try:
        if os.path.isdir(share_path):
            logger.info(f"scan_share_directory: '{share_path}' is a valid directory. Scanning...")
            files = os.listdir(share_path)
            logger.info(f"scan_share_directory: Found {len(files)} items.")
            for entry in files:
                lower_entry = entry.lower()
                if (lower_entry.endswith('.xlsx') or lower_entry.endswith('.xls')) and not entry.startswith('~$'):
                    filepath = os.path.join(share_path, entry)
                    dt = parse_filename_timestamp(entry, filepath)
                    results.append((entry, filepath, dt))
            # Sort by timestamp, latest first
            results.sort(key=lambda x: x[2], reverse=True)
            logger.info(f"scan_share_directory: Returning {len(results)} valid excel files.")
        else:
            logger.warning(f"scan_share_directory: '{share_path}' is NOT a directory or not accessible.")
    except Exception as e:
        logger.error(f"Error scanning share: {e}", exc_info=True)
    return results

def get_loaded_files(db_path):
    """Return list of all loaded file records."""
    conn = sqlite3.connect(db_path)
    conn.row_factory = _dict_factory
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM loaded_files ORDER BY loaded_at DESC")
    rows = cursor.fetchall()
    conn.close()
    return rows

def get_last_fetch_time(db_path):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT MAX(loaded_at) FROM loaded_files WHERE source = 'network_share'")
    result = cursor.fetchone()
    conn.close()
    if result and result[0]:
        try:
            return datetime.fromisoformat(result[0])
        except ValueError:
            return None
    return None

def _extract_sn(inventory_str):
    """Extracts serial number from inventory string and determines if it is XXX-like."""
    if not inventory_str:
        return "", 1
    
    # Example: 'VALVE (PN: 3216000-00, SN: 3216000-00311)'
    sn_match = re.search(r'SN:\s*([^)]+)', inventory_str)
    sn = sn_match.group(1).strip() if sn_match else ""
    
    xxx_values = ['XXX', 'UNKNOWN', 'N/A', 'NONE', '']
    has_xxx_sn = 1 if sn.upper() in xxx_values else 0
    
    return sn, has_xxx_sn

def load_excel(db_path, excel_path):
    """Loads data from the given Excel file into the database."""
    logger.info(f"Loading data from {excel_path} into {db_path}")
    
    rows_iter = None
    if excel_path.lower().endswith('.xls'):
        try:
            import xlrd
            workbook = xlrd.open_workbook(excel_path)
            sheet_name = 'Export Worksheet' if 'Export Worksheet' in workbook.sheet_names() else workbook.sheet_names()[0]
            sheet = workbook.sheet_by_name(sheet_name)
            
            def xlrd_rows():
                for row_idx in range(sheet.nrows):
                    row_vals = []
                    for col_idx in range(sheet.ncols):
                        cell_type = sheet.cell_type(row_idx, col_idx)
                        cell_value = sheet.cell_value(row_idx, col_idx)
                        if cell_type == xlrd.XL_CELL_DATE:
                            try:
                                dt = xlrd.xldate_as_datetime(cell_value, workbook.datemode)
                                row_vals.append(dt)
                            except Exception:
                                row_vals.append(cell_value)
                        else:
                            row_vals.append(cell_value)
                    yield row_vals
            
            rows_iter = xlrd_rows()
            logger.info(f"Using sheet: {sheet_name} (xlrd)")
        except ImportError:
            logger.error("xlrd library is not installed for reading .xls files.")
            return False, 0
        except Exception as e:
            logger.error(f"Failed to open .xls file: {e}")
            return False, 0
    else:
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            # Try 'Export Worksheet' first, fall back to first sheet
            if 'Export Worksheet' in wb.sheetnames:
                sheet = wb['Export Worksheet']
            else:
                sheet = wb[wb.sheetnames[0]]
                logger.info(f"Using sheet: {wb.sheetnames[0]}")
                
            def openpyxl_rows():
                for row in sheet.rows:
                    yield [cell.value for cell in row]
            
            rows_iter = openpyxl_rows()
        except Exception as e:
            logger.error(f"Failed to open Excel file: {e}")
            return False, 0
    
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Header check
    try:
        header = next(rows_iter)
    except StopIteration:
        logger.error("Empty sheet")
        return False, 0
        
    inserted_count = 0
    for values in rows_iter:
        # Skip empty rows completely
        if not any(values):
            continue
        # Skip header-like rows (e.g. from .xls files where header wasn't consumed)
        if values[0] and str(values[0]).strip().upper() == 'ASSMBL_CD':
            continue
            
        def safe_str(val):
            return str(val) if val is not None else ""
            
        assembly_cd = safe_str(values[0]) if len(values) > 0 else ""
        nh_assembly_cd = safe_str(values[1]) if len(values) > 1 else ""
        bom_class_cd = safe_str(values[2]) if len(values) > 2 else ""
        config_slot_code = safe_str(values[3]) if len(values) > 3 else ""
        config_slot_name = safe_str(values[4]) if len(values) > 4 else ""
        part_group_cd = safe_str(values[5]) if len(values) > 5 else ""
        part_group_name = safe_str(values[6]) if len(values) > 6 else ""
        
        event_dt_raw = values[7] if len(values) > 7 else None
        event_dt = ""
        if isinstance(event_dt_raw, datetime):
            event_dt = event_dt_raw.isoformat()
        elif event_dt_raw:
            event_dt = safe_str(event_dt_raw)
            
        event_type = safe_str(values[8]) if len(values) > 8 else ""
        status_cd = safe_str(values[9]) if len(values) > 9 else ""
        event_desc = safe_str(values[10]) if len(values) > 10 else ""
        aircraft = safe_str(values[11]) if len(values) > 11 else ""
        inventory_key = safe_str(values[12]) if len(values) > 12 else ""
        inventory = safe_str(values[13]) if len(values) > 13 else ""
        barcode = safe_str(values[14]) if len(values) > 14 else ""
        config_slot = safe_str(values[15]) if len(values) > 15 else ""
        part_no = safe_str(values[16]) if len(values) > 16 else ""
        part_desc = safe_str(values[17]) if len(values) > 17 else ""
        remove_reason = safe_str(values[18]) if len(values) > 18 else ""
        performed_by_user = safe_str(values[19]) if len(values) > 19 else ""
        performed_by_username = safe_str(values[20]) if len(values) > 20 else ""
        performed_by_hr_cd = safe_str(values[21]) if len(values) > 21 else ""
        
        sn, has_xxx_sn = _extract_sn(inventory)
        
        try:
            cursor.execute('''
            INSERT INTO htc_events (
                assembly_cd, nh_assembly_cd, bom_class_cd, config_slot_code, config_slot_name,
                part_group_cd, part_group_name, event_dt, event_type, status_cd,
                event_desc, aircraft, inventory_key, inventory, barcode,
                config_slot, part_no, part_desc, remove_reason, performed_by_user,
                performed_by_username, performed_by_hr_cd, serial_number, has_xxx_sn
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                assembly_cd, nh_assembly_cd, bom_class_cd, config_slot_code, config_slot_name,
                part_group_cd, part_group_name, event_dt, event_type, status_cd,
                event_desc, aircraft, inventory_key, inventory, barcode,
                config_slot, part_no, part_desc, remove_reason, performed_by_user,
                performed_by_username, performed_by_hr_cd, sn, has_xxx_sn
            ))
            inserted_count += 1
        except sqlite3.IntegrityError:
            # Skip duplicates based on unique constraint
            pass
            
    conn.commit()
    conn.close()
    return True, inserted_count

def _dict_factory(cursor, row):
    d = {}
    for idx, col in enumerate(cursor.description):
        d[col[0]] = row[idx]
    return d

def get_events(db_path, fleet=None, aircraft=None, event_type=None, date_from=None, date_to=None, search=None, bom=None, part_group=None, page=1, per_page=50):
    conn = sqlite3.connect(db_path)
    conn.row_factory = _dict_factory
    cursor = conn.cursor()
    
    query = "SELECT * FROM htc_events WHERE 1=1"
    params = []
    
    if fleet:
        query += " AND assembly_cd = ?"
        params.append(fleet)
    if aircraft:
        query += " AND aircraft LIKE ?"
        params.append(f"%{aircraft}%")
    if event_type:
        query += " AND event_type = ?"
        params.append(event_type)
    if date_from:
        query += " AND event_dt >= ?"
        params.append(date_from)
    if date_to:
        query += " AND event_dt <= ?"
        params.append(date_to + 'T23:59:59' if len(date_to) == 10 else date_to)
    if bom:
        query += " AND config_slot_code LIKE ?"
        params.append(f"%{bom}%")
    if part_group:
        query += " AND part_group_name LIKE ?"
        params.append(f"%{part_group}%")
    if search:
        query += " AND (barcode LIKE ? OR part_no LIKE ? OR serial_number LIKE ?)"
        params.extend([f"%{search}%", f"%{search}%", f"%{search}%"])
        
    # Count total
    count_query = f"SELECT COUNT(*) as total FROM ({query})"
    cursor.execute(count_query, params)
    total = cursor.fetchone()['total']
    
    # Pagination
    offset = (page - 1) * per_page
    query += " ORDER BY id DESC LIMIT ? OFFSET ?"
    params.extend([per_page, offset])
    
    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    
    return {
        "data": rows,
        "total": total,
        "page": page,
        "per_page": per_page
    }

def get_total_event_count(db_path):
    """Cheap existence/count check — used at startup instead of get_dashboard_stats,
    which builds full result sets across three queries just to check for emptiness."""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM htc_events")
    count = cursor.fetchone()[0]
    conn.close()
    return count


def get_dashboard_stats(db_path, fleet=None, aircraft=None, date_from=None, date_to=None, bom=None, part_group=None):
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    stats = {
        "total_events": 0,
        "install_count": 0,
        "remove_count": 0,
        "xxx_sn_count": 0,
        "empty_slots_count": 0,
        "fleet_breakdown": {}
    }
    
    all_events = get_alert_a_events(db_path, fleet=fleet, aircraft=aircraft, date_from=date_from, date_to=date_to, bom=bom, part_group=part_group)
    stats["total_events"] = len(all_events)
    stats["install_count"] = sum(1 for e in all_events if e['event_type'] == 'INSTALL')
    stats["remove_count"] = sum(1 for e in all_events if e['event_type'] == 'REMOVE')
    
    alert_b = get_alert_b_events(db_path, fleet=fleet, aircraft=aircraft, date_from=date_from, date_to=date_to, bom=bom, part_group=part_group)
    stats["xxx_sn_count"] = len(alert_b)
    
    mmc_alerts = get_mmc_alerts(db_path, fleet=fleet, aircraft=aircraft, date_from=date_from, date_to=date_to, bom=bom, part_group=part_group)
    stats["empty_slots_count"] = len(mmc_alerts)
    
    cursor.execute("SELECT assembly_cd, COUNT(*) FROM htc_events GROUP BY assembly_cd")
    for row in cursor.fetchall():
        if row[0]:
            stats["fleet_breakdown"][row[0]] = row[1]
            
    conn.close()
    return stats

def get_alert_a_events(db_path, fleet=None, aircraft=None, event_type=None, date_from=None, date_to=None, bom=None, part_group=None):
    conn = sqlite3.connect(db_path)
    conn.row_factory = _dict_factory
    cursor = conn.cursor()
    
    query = "SELECT * FROM htc_events WHERE 1=1"
    params = []
    
    if event_type and event_type in ['INSTALL', 'REMOVE']:
        query += " AND event_type = ?"
        params.append(event_type)
    else:
        query += " AND event_type IN ('INSTALL', 'REMOVE')"
        
    if fleet:
        query += " AND assembly_cd = ?"
        params.append(fleet)
    if aircraft:
        query += " AND aircraft LIKE ?"
        params.append(f"%{aircraft}%")
    if date_from:
        query += " AND event_dt >= ?"
        params.append(date_from)
    if date_to:
        query += " AND event_dt <= ?"
        params.append(date_to + 'T23:59:59' if len(date_to) == 10 else date_to)
    if bom:
        query += " AND config_slot_code LIKE ?"
        params.append(f"%{bom}%")
    if part_group:
        query += " AND part_group_name LIKE ?"
        params.append(f"%{part_group}%")
        
    query += " ORDER BY event_dt DESC"
    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    return rows

def get_alert_b_events(db_path, fleet=None, aircraft=None, date_from=None, date_to=None, bom=None, part_group=None):
    conn = sqlite3.connect(db_path)
    conn.row_factory = _dict_factory
    cursor = conn.cursor()
    
    query = """
    SELECT e1.* FROM htc_events e1
    WHERE e1.has_xxx_sn = 1
    AND e1.event_type = 'INSTALL'
    AND NOT EXISTS (
        SELECT 1 FROM htc_events e2 
        WHERE e2.aircraft = e1.aircraft 
          AND e2.config_slot_code = e1.config_slot_code 
          AND e2.config_slot = e1.config_slot 
          AND (e2.event_dt > e1.event_dt OR (e2.event_dt = e1.event_dt AND e2.id > e1.id))
    )
    """
    params = []
    
    if fleet:
        query += " AND e1.assembly_cd = ?"
        params.append(fleet)
    if aircraft:
        query += " AND e1.aircraft LIKE ?"
        params.append(f"%{aircraft}%")
    if date_from:
        query += " AND e1.event_dt >= ?"
        params.append(date_from)
    if date_to:
        query += " AND e1.event_dt <= ?"
        params.append(date_to + 'T23:59:59' if len(date_to) == 10 else date_to)
    if bom:
        query += " AND e1.config_slot_code LIKE ?"
        params.append(f"%{bom}%")
    if part_group:
        query += " AND e1.part_group_name LIKE ?"
        params.append(f"%{part_group}%")
        
    query += " GROUP BY e1.config_slot, e1.status_cd, e1.inventory_key, e1.performed_by_user, e1.performed_by_username ORDER BY e1.id DESC"
    
    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    return rows


def _parse_event_date(dt_str):
    """Parse event date strings in multiple formats. Returns datetime or None."""
    if not dt_str:
        return None
    # Try DD-MON-YYYY HH:MI:SS (e.g. '02-JUL-2026 15:54:34')
    try:
        return datetime.strptime(dt_str, '%d-%b-%Y %H:%M:%S')
    except (ValueError, TypeError):
        pass
    # Try ISO format (e.g. '2026-07-02T15:54:34')
    try:
        return datetime.fromisoformat(dt_str)
    except (ValueError, TypeError):
        pass
    # Try DD-MON-YYYY (no time)
    try:
        return datetime.strptime(dt_str, '%d-%b-%Y')
    except (ValueError, TypeError):
        pass
    return None


def _get_inv_key_num(inv_key_str):
    if not inv_key_str:
        return 0
    try:
        parts = inv_key_str.split(':')
        if len(parts) > 1:
            return int(parts[1])
        return int(parts[0])
    except:
        return 0

def get_mmc_alerts(db_path, fleet=None, aircraft=None, date_from=None, date_to=None, bom=None, part_group=None):
    """Detect Missing Mandatory Component (MMC) alerts.
    
    A removed part must be reinstalled on the same aircraft + config slot
    within 7 days. If no matching INSTALL is found within that window,
    the removal is flagged as MMC.
    """
    conn = sqlite3.connect(db_path)
    conn.row_factory = _dict_factory
    cursor = conn.cursor()
    
    # Get all REMOVE events
    remove_query = "SELECT * FROM htc_events WHERE event_type = 'REMOVE'"
    params = []
    if fleet:
        remove_query += " AND assembly_cd = ?"
        params.append(fleet)
    if aircraft:
        remove_query += " AND aircraft LIKE ?"
        params.append(f"%{aircraft}%")
    if date_from:
        remove_query += " AND event_dt >= ?"
        params.append(date_from)
    if date_to:
        remove_query += " AND event_dt <= ?"
        params.append(date_to + 'T23:59:59' if len(date_to) == 10 else date_to)
    if bom:
        remove_query += " AND config_slot_code LIKE ?"
        params.append(f"%{bom}%")
    if part_group:
        remove_query += " AND part_group_name LIKE ?"
        params.append(f"%{part_group}%")
    remove_query += " ORDER BY id DESC"
    
    cursor.execute(remove_query, params)
    removals_raw = cursor.fetchall()
    
    # Get all INSTALL events for matching
    cursor.execute("SELECT * FROM htc_events WHERE event_type = 'INSTALL'")
    installs_raw = cursor.fetchall()
    conn.close()
    
    # Parse dates and filter valid ones
    removals = []
    for r in removals_raw:
        dt = _parse_event_date(r.get('event_dt'))
        if dt:
            removals.append((dt, r))
            
    installs = []
    for i in installs_raw:
        dt = _parse_event_date(i.get('event_dt'))
        if dt:
            installs.append((dt, i))
            
    # Sort chronologically (by date, then FG_INVENTORY_KEY as tie-breaker)
    removals.sort(key=lambda x: (x[0], _get_inv_key_num(x[1].get('inventory_key', ''))))
    installs.sort(key=lambda x: (x[0], _get_inv_key_num(x[1].get('inventory_key', ''))))
    
    now = datetime.now()
    mmc_alerts = []
    claimed_installs = set()
    
    for removal_dt, removal in removals:
        aircraft = removal.get('aircraft', '')
        config_slot_code = removal.get('config_slot_code', '')
        config_slot = removal.get('config_slot', '')
        rem_inv_num = _get_inv_key_num(removal.get('inventory_key', ''))
        
        # Find the earliest INSTALL on the same aircraft + config_slot_code + config_slot that hasn't been claimed
        matching_install = None
        for inst_dt, inst in installs:
            inst_id = inst.get('id')
            if inst_id in claimed_installs:
                continue
                
            if (inst.get('aircraft', '') == aircraft and 
                inst.get('config_slot_code', '') == config_slot_code and
                inst.get('config_slot', '') == config_slot and
                inst.get('has_xxx_sn') == 0):
                
                inst_inv_num = _get_inv_key_num(inst.get('inventory_key', ''))
                
                # Match if install is strictly after in time
                # OR if time is identical, the install's inventory_key is GREATER THAN OR EQUAL to removal's
                if inst_dt > removal_dt or (inst_dt == removal_dt and inst_inv_num >= rem_inv_num):
                    matching_install = inst
                    claimed_installs.add(inst_id)
                    break  # earliest matching install
        
        days_since = (now - removal_dt).days
        
        if matching_install:
            # A valid SN was installed; the MMC issue is resolved and closed
            continue
        else:
            # No valid reinstall found at all - remains open
            if days_since >= 3:
                alert = dict(removal)
                alert['days_since_removal'] = days_since
                alert['days_to_reinstall'] = None
                alert['reinstall_date'] = None
                alert['reinstall_sn'] = None
                alert['_parsed_dt'] = removal_dt
                if days_since >= 7:
                    alert['mmc_severity'] = 'CRITICAL'
                    alert['reinstall_status'] = 'OVERDUE'
                else:
                    alert['mmc_severity'] = 'WARNING'
                    alert['reinstall_status'] = 'PENDING'
                mmc_alerts.append(alert)
                
    # Sort newest first for the UI
    mmc_alerts.sort(key=lambda x: x['_parsed_dt'], reverse=True)
    for a in mmc_alerts:
        del a['_parsed_dt']
    
    return mmc_alerts

def get_fleet_dashboard(db_path, aircraft=None, date_from=None, date_to=None, bom=None, part_group=None):
    """Returns XXX S/N Alert counts and Empty Config Slot (MMC) counts, broken
    down per fleet type, for the Dashboard tab's per-fleet KPI cards."""
    fleets = get_fleet_types(db_path)
    
    breakdown = []
    for fleet in fleets:
        if not fleet: continue
        alert_b = get_alert_b_events(db_path, fleet=fleet, aircraft=aircraft, date_from=date_from, date_to=date_to, bom=bom, part_group=part_group)
        mmc = get_mmc_alerts(db_path, fleet=fleet, aircraft=aircraft, date_from=date_from, date_to=date_to, bom=bom, part_group=part_group)
        breakdown.append({
            "fleet": fleet,
            "xxx_sn_count": len(alert_b),
            "empty_slots_count": len(mmc)
        })

    # Sort fleets with the most XXX S/N alerts first so the busiest fleets surface at the top
    breakdown.sort(key=lambda x: x["xxx_sn_count"], reverse=True)
    return breakdown


def get_fleet_types(db_path):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT assembly_cd FROM htc_events WHERE assembly_cd IS NOT NULL AND assembly_cd != ''")
    rows = cursor.fetchall()
    conn.close()
    return [r[0] for r in rows]

def get_empty_slots(db_path):
    conn = sqlite3.connect(db_path)
    conn.row_factory = _dict_factory
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM htc_events WHERE event_dt IS NULL OR event_dt = '' ORDER BY id DESC")
    rows = cursor.fetchall()
    conn.close()
    return rows

def get_resolutions(db_path):
    conn = sqlite3.connect(db_path)
    conn.row_factory = _dict_factory
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM resolutions ORDER BY id DESC")
    rows = cursor.fetchall()
    conn.close()
    return rows

def create_resolution(db_path, data):
    conn = sqlite3.connect(db_path)
    conn.row_factory = _dict_factory
    cursor = conn.cursor()
    
    now = datetime.utcnow().isoformat()
    cursor.execute('''
    INSERT INTO resolutions (
        event_barcode, alert_date, aircraft, config_slot, part_no,
        original_sn, resolved_sn, engineer_responsible, resolution_date,
        status, notes, created_at, updated_at
    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        data.get('event_barcode'),
        data.get('alert_date'),
        data.get('aircraft'),
        data.get('config_slot'),
        data.get('part_no'),
        data.get('original_sn'),
        data.get('resolved_sn'),
        data.get('engineer_responsible'),
        data.get('resolution_date'),
        data.get('status', 'PENDING'),
        data.get('notes'),
        now,
        now
    ))
    conn.commit()
    res_id = cursor.lastrowid
    
    cursor.execute("SELECT * FROM resolutions WHERE id = ?", (res_id,))
    row = cursor.fetchone()
    conn.close()
    return row

def update_resolution(db_path, res_id, data):
    conn = sqlite3.connect(db_path)
    conn.row_factory = _dict_factory
    cursor = conn.cursor()
    
    now = datetime.utcnow().isoformat()
    
    set_clauses = []
    params = []
    
    for key, value in data.items():
        if key in ['resolved_sn', 'engineer_responsible', 'resolution_date', 'status', 'notes']:
            set_clauses.append(f"{key} = ?")
            params.append(value)
            
    if set_clauses:
        set_clauses.append("updated_at = ?")
        params.append(now)
        
        query = f"UPDATE resolutions SET {', '.join(set_clauses)} WHERE id = ?"
        params.append(res_id)
        
        cursor.execute(query, params)
        conn.commit()
        
    cursor.execute("SELECT * FROM resolutions WHERE id = ?", (res_id,))
    row = cursor.fetchone()
    conn.close()
    return row

def get_bom_types(db_path):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT config_slot_code FROM htc_events WHERE config_slot_code IS NOT NULL AND config_slot_code != '' ORDER BY config_slot_code")
    boms = [row[0] for row in cursor.fetchall()]
    conn.close()
    return boms

def get_part_groups(db_path):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT part_group_name FROM htc_events WHERE part_group_name IS NOT NULL AND part_group_name != '' ORDER BY part_group_name")
    groups = [row[0] for row in cursor.fetchall()]
    conn.close()
    return groups


# ─────────────────────────────────────────────────────────────────────────────
# Email subscriptions (HTC Visibility Monitoring digest emails)
# ─────────────────────────────────────────────────────────────────────────────

_VALID_ALERT_TYPES = {"ALERT_B", "ALERT_C"}
_VALID_FREQUENCIES = {"every_3_days", "weekly"}
_VALID_DAYS = {"mon", "tue", "wed", "thu", "fri", "sat", "sun"}


def get_subscriptions(db_path, active_only=False):
    conn = sqlite3.connect(db_path)
    conn.row_factory = _dict_factory
    cursor = conn.cursor()
    query = "SELECT * FROM email_subscriptions"
    if active_only:
        query += " WHERE active = 1"
    query += " ORDER BY id DESC"
    cursor.execute(query)
    rows = cursor.fetchall()
    conn.close()
    return [_expand_fleet_types(_expand_alert_types(r)) for r in rows]


def get_subscription(db_path, sub_id):
    conn = sqlite3.connect(db_path)
    conn.row_factory = _dict_factory
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM email_subscriptions WHERE id = ?", (sub_id,))
    row = cursor.fetchone()
    conn.close()
    return _expand_fleet_types(_expand_alert_types(row)) if row else row


def _serialize_fleet_types(fleet_type_input):
    """Accepts either a list of fleet types or a single string/comma-separated
    string, dedupes while preserving order, and returns a comma-separated
    string for storage. Fleet types are open-ended (sourced from event data),
    so unlike alert types there's no fixed allow-list to validate against."""
    if fleet_type_input is None:
        raw_list = []
    elif isinstance(fleet_type_input, (list, tuple, set)):
        raw_list = list(fleet_type_input)
    else:
        raw_list = str(fleet_type_input).split(",")

    seen = set()
    cleaned = []
    for item in raw_list:
        val = str(item).strip()
        if not val:
            continue
        if val not in seen:
            seen.add(val)
            cleaned.append(val)

    if not cleaned:
        raise ValueError("At least one fleet_type is required.")

    return ",".join(cleaned)


def _expand_fleet_types(row):
    """Adds a 'fleet_types' list field (parsed from the stored comma-separated
    'fleet_type' string) onto a subscription row dict for easier frontend use,
    while keeping 'fleet_type' as the raw stored string for backward compat."""
    if row and row.get("fleet_type"):
        row["fleet_types"] = [f for f in row["fleet_type"].split(",") if f]
    elif row is not None:
        row["fleet_types"] = []
    return row


def _serialize_alert_types(alert_type_input):
    """Accepts either a list of alert types or a single string/comma-separated
    string, validates each against _VALID_ALERT_TYPES, dedupes while
    preserving order, and returns a comma-separated string for storage."""
    if alert_type_input is None:
        raw_list = []
    elif isinstance(alert_type_input, (list, tuple, set)):
        raw_list = list(alert_type_input)
    else:
        raw_list = str(alert_type_input).split(",")

    seen = set()
    cleaned = []
    for item in raw_list:
        val = str(item).strip().upper()
        if not val:
            continue
        if val not in _VALID_ALERT_TYPES:
            raise ValueError(f"alert_type must be one or more of: {', '.join(sorted(_VALID_ALERT_TYPES))}. Got '{val}'.")
        if val not in seen:
            seen.add(val)
            cleaned.append(val)

    if not cleaned:
        raise ValueError(f"At least one alert_type is required (one or more of: {', '.join(sorted(_VALID_ALERT_TYPES))}).")

    return ",".join(cleaned)


def _expand_alert_types(row):
    """Adds an 'alert_types' list field (parsed from the stored comma-separated
    'alert_type' string) onto a subscription row dict for easier frontend use,
    while keeping 'alert_type' as the raw stored string for backward compat."""
    if row and row.get("alert_type"):
        row["alert_types"] = [a for a in row["alert_type"].split(",") if a]
    elif row is not None:
        row["alert_types"] = []
    return row


def create_subscription(db_path, data):
    """Create a new email subscription.

    Required keys: fleet_type (list of one or more fleet types, or a
    comma-separated string — also accepts the key 'fleet_types' as an alias),
    alert_type (list of one or more of 'ALERT_B'/'ALERT_C', or a
    comma-separated string — also accepts the key 'alert_types' as an alias),
    frequency ('every_3_days'|'weekly'), run_time ('HH:MM'), email.
    day_of_week required when frequency == 'weekly'.

    Returns the created row dict, or raises ValueError on invalid input.
    """
    fleet_type_input = data.get("fleet_type", data.get("fleet_types"))
    alert_type_input = data.get("alert_type", data.get("alert_types"))
    frequency = (data.get("frequency") or "").strip().lower()
    day_of_week = (data.get("day_of_week") or "").strip().lower() or None
    run_time = (data.get("run_time") or "").strip()
    email = (data.get("email") or "").strip().lower()

    fleet_type = _serialize_fleet_types(fleet_type_input)
    alert_type = _serialize_alert_types(alert_type_input)
    if frequency not in _VALID_FREQUENCIES:
        raise ValueError("frequency must be 'every_3_days' or 'weekly'.")
    if frequency == "weekly" and day_of_week not in _VALID_DAYS:
        raise ValueError("day_of_week is required and must be a 3-letter day (mon..sun) for weekly frequency.")
    if not re.match(r'^\d{2}:\d{2}$', run_time):
        raise ValueError("run_time must be in 'HH:MM' 24-hour format.")
    if not email or "@" not in email:
        raise ValueError("A valid email address is required.")

    conn = sqlite3.connect(db_path)
    conn.row_factory = _dict_factory
    cursor = conn.cursor()
    now = datetime.utcnow().isoformat()
    cursor.execute('''
        INSERT INTO email_subscriptions (
            fleet_type, alert_type, frequency, day_of_week, run_time,
            email, active, created_at, updated_at, last_sent_at
        ) VALUES (?, ?, ?, ?, ?, ?, 1, ?, ?, NULL)
    ''', (fleet_type, alert_type, frequency, day_of_week, run_time, email, now, now))
    conn.commit()
    sub_id = cursor.lastrowid
    cursor.execute("SELECT * FROM email_subscriptions WHERE id = ?", (sub_id,))
    row = cursor.fetchone()
    conn.close()
    return _expand_fleet_types(_expand_alert_types(row))


def update_subscription(db_path, sub_id, data):
    """Update an existing subscription. Any of the create_subscription fields
    may be supplied; only supplied fields are validated and updated."""
    allowed_fields = {
        "fleet_type", "alert_type", "frequency", "day_of_week",
        "run_time", "email", "active",
    }
    set_clauses = []
    params = []

    if "fleet_type" in data or "fleet_types" in data:
        val = _serialize_fleet_types(data.get("fleet_type", data.get("fleet_types")))
        set_clauses.append("fleet_type = ?")
        params.append(val)

    if "alert_type" in data or "alert_types" in data:
        val = _serialize_alert_types(data.get("alert_type", data.get("alert_types")))
        set_clauses.append("alert_type = ?")
        params.append(val)

    if "frequency" in data:
        val = (data.get("frequency") or "").strip().lower()
        if val not in _VALID_FREQUENCIES:
            raise ValueError("frequency must be 'every_3_days' or 'weekly'.")
        set_clauses.append("frequency = ?")
        params.append(val)

    if "day_of_week" in data:
        val = (data.get("day_of_week") or "").strip().lower() or None
        if val is not None and val not in _VALID_DAYS:
            raise ValueError("day_of_week must be a 3-letter day (mon..sun).")
        set_clauses.append("day_of_week = ?")
        params.append(val)

    if "run_time" in data:
        val = (data.get("run_time") or "").strip()
        if not re.match(r'^\d{2}:\d{2}$', val):
            raise ValueError("run_time must be in 'HH:MM' 24-hour format.")
        set_clauses.append("run_time = ?")
        params.append(val)

    if "email" in data:
        val = (data.get("email") or "").strip().lower()
        if not val or "@" not in val:
            raise ValueError("A valid email address is required.")
        set_clauses.append("email = ?")
        params.append(val)

    if "active" in data:
        set_clauses.append("active = ?")
        params.append(1 if data.get("active") else 0)

    if not set_clauses:
        # Nothing to update — just return current row
        return get_subscription(db_path, sub_id)

    conn = sqlite3.connect(db_path)
    conn.row_factory = _dict_factory
    cursor = conn.cursor()

    now = datetime.utcnow().isoformat()
    set_clauses.append("updated_at = ?")
    params.append(now)

    query = f"UPDATE email_subscriptions SET {', '.join(set_clauses)} WHERE id = ?"
    params.append(sub_id)
    cursor.execute(query, params)
    conn.commit()

    cursor.execute("SELECT * FROM email_subscriptions WHERE id = ?", (sub_id,))
    row = cursor.fetchone()
    conn.close()
    return _expand_fleet_types(_expand_alert_types(row))


def delete_subscription(db_path, sub_id):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM email_subscriptions WHERE id = ?", (sub_id,))
    conn.commit()
    deleted = cursor.rowcount > 0
    conn.close()
    return deleted


def mark_subscription_sent(db_path, sub_id, when=None):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    ts = (when or datetime.utcnow()).isoformat()
    cursor.execute(
        "UPDATE email_subscriptions SET last_sent_at = ? WHERE id = ?",
        (ts, sub_id)
    )
    conn.commit()
    conn.close()


# ─────────────────────────────────────────────────────────────────────────────
# Generic app config (key/value) — used for OWA password, admin password, etc.
# ─────────────────────────────────────────────────────────────────────────────

def get_config_value(db_path, key, default=None):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT value FROM app_config WHERE key = ?", (key,))
    row = cursor.fetchone()
    conn.close()
    if row and row[0] is not None:
        return row[0]
    return default


def set_config_value(db_path, key, value):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO app_config (key, value) VALUES (?, ?)
        ON CONFLICT(key) DO UPDATE SET value = excluded.value
    ''', (key, value))
    conn.commit()
    conn.close()